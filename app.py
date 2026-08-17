"""
H2 2026 Command Center - PODS 2
Deal Execution Tracker & Strategy Dashboard (SQLite-backed version)

Row-level access control: each Account Manager can only edit the opportunities
assigned to them; ADMIN can edit everything; MANAGEMENT is read-only.
"""
import io
import json
import os
import secrets
import sqlite3
from datetime import datetime, date
from functools import wraps

from flask import Flask, g, jsonify, render_template, request, send_file
from werkzeug.security import check_password_hash, generate_password_hash

# Keep the database next to app.py so it persists in a predictable location
# regardless of the host's working directory (Render, PythonAnywhere, Docker, etc.).
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.environ.get("DB_PATH", os.path.join(BASE_DIR, "db.sqlite3"))

app = Flask(__name__)

# In-memory token store: token -> {user_id, username, role, full_name}
TOKENS = {}

VALID_ROLES = ("admin", "account_manager", "management")
VALID_STAGES = ("Prospecting", "Negotiation", "Closed", "Blocked")

DEFAULT_PILLARS = [
    "IoT Connectivity",
    "Device Bundling",
    "CCTV & Vision Analytics",
    "Enterprise Solutions",
    "Digital Reward",
]
DEFAULT_SQUADS = ["Volume Squad", "Tender Squad", "Strategic Squad"]
DEFAULT_STAGES = ["Prospecting", "Negotiation", "Closed", "Blocked"]

# --------------------------------------------------------------------------
# Execution Framework - the 8 Enterprise Proofs (PODS 2 execution standard)
# --------------------------------------------------------------------------
PROOFS = [
    ("qualification", "Proof of Qualification",
     "Ensure target customers (B500) have the need, budget, and alignment with our solution"),
    ("engagement", "Proof of Engagement",
     "Secure deep engagement with all key customer stakeholders (PODS, SPIN selling, Sales card)"),
    ("concept", "Proof of Concept",
     "Verify the proposed solution functions exactly as specified (TQC)"),
    ("value", "Proof of Value",
     "Confirm the solution delivers a profitable outcome for the customer (TQC & outcome-based solution)"),
    ("contract", "Proof of Contract",
     "Ensure all TCLRF terms are covered (Technical, Commercial, Legal, Risk, Framework)"),
    ("delivery", "Proof of Delivery",
     "Ensure the final solution is delivered according to all expectations (TQC)"),
    ("operation", "Proof of Operation",
     "Commit on agreed-upon Service Level Agreement (SLA) and ensure payment"),
    ("clm", "Proof of CLM",
     "Nurture the customer relationship to prevent churn and enable upselling (Customer card)"),
]
PROOF_KEYS = [p[0] for p in PROOFS]
PROOF_NAMES = {p[0]: p[1] for p in PROOFS}
PROOF_STATUSES = ("not_started", "in_progress", "done", "na")
ENTRY_STATUSES = ("not_started", "planned", "in_progress", "done")
ENTRY_STATUS_LABELS = {
    "not_started": "Not started",
    "planned": "Planned",
    "in_progress": "In progress",
    "done": "Done",
}


def normalize_proofs(raw):
    """Always return the full 8-proof structure, preserving whatever was stored.

    Each proof holds a list of evidence `entries` ({text, date, status}). A legacy
    single `notes` string is migrated into the first entry so nothing is ever lost.
    Entries without their own status fall back to the parent proof's status
    (an "na" proof falls back to "not_started") so older data keeps working.
    """
    incoming = raw if isinstance(raw, dict) else {}
    out = {}
    for key in PROOF_KEYS:
        item = incoming.get(key) or {}
        if not isinstance(item, dict):
            item = {}
        status = str(item.get("status", "not_started")).strip().lower().replace(" ", "_")
        if status not in PROOF_STATUSES:
            status = "not_started"
        fallback_entry_status = status if status in ENTRY_STATUSES else "not_started"

        entries = []
        for e in (item.get("entries") or []):
            if isinstance(e, dict):
                text = str(e.get("text", "") or "").strip()
                date = str(e.get("date", "") or "").strip()[:10]
                estatus = str(e.get("status", "") or "").strip().lower().replace(" ", "_")
                if estatus not in ENTRY_STATUSES:
                    estatus = fallback_entry_status
            else:
                text, date, estatus = str(e or "").strip(), "", fallback_entry_status
            if text:
                entries.append({"text": text, "date": date, "status": estatus})

        # Migrate a legacy notes string into the evidence list.
        legacy = str(item.get("notes", "") or "").strip()
        if legacy and not any(e["text"] == legacy for e in entries):
            entries.insert(0, {"text": legacy, "date": "", "status": fallback_entry_status})

        out[key] = {
            "status": status,
            "due": str(item.get("due", "") or "")[:10],
            "entries": entries,
            # kept for backward compatibility with older clients/exports
            "notes": entries[0]["text"] if entries else "",
        }
    return out


def proof_progress(proofs):
    """% of applicable (non-N/A) proofs completed."""
    p = normalize_proofs(proofs)
    applicable = [v for v in p.values() if v["status"] != "na"]
    if not applicable:
        return 0
    done = sum(1 for v in applicable if v["status"] == "done")
    return round(done / len(applicable) * 100)


# Which stage each proof implies once it is reached. Admin-editable in Configuration.
DEFAULT_STAGE_RULES = {
    "qualification": "Prospecting",
    "engagement": "Prospecting",
    "concept": "Negotiation",
    "value": "Negotiation",
    "contract": "Negotiation",
    "delivery": "Closed",
    "operation": "Closed",
    "clm": "Closed",
}
DEFAULT_BLOCKED_STAGE = "Blocked"


def derive_stage(proofs, is_blocked, rules, blocked_stage, fallback):
    """Stage implied by how far the execution framework has progressed.

    A blocked deal always shows the blocked stage. Otherwise we take the furthest
    proof that has been reached (done or in progress) and use its mapped stage.
    """
    if is_blocked and blocked_stage:
        return blocked_stage
    p = normalize_proofs(proofs)
    reached = None
    for key in PROOF_KEYS:                      # PROOF_KEYS is in framework order
        if p[key]["status"] in ("done", "in_progress"):
            reached = key
    if reached is None:
        return fallback
    return (rules or {}).get(reached) or fallback

# Seed Account Managers: (username, password, full_name)
SEED_AMS = [
    ("anisa", "anisa123", "Anisa Rahmy"),
    ("arie", "arie123", "Arie Prabowo"),
    ("ashari", "ashari123", "Ashari"),
    ("dimas", "dimas123", "Dimas"),
]

# Example per-AM revenue targets for H2 2026 (admin can edit in Settings).
DEFAULT_AM_TARGETS = {
    "Anisa Rahmy": 15_000_000_000,
    "Arie Prabowo": 13_000_000_000,
    "Ashari": 14_000_000_000,
    "Dimas": 15_000_000_000,
}


def _hash(pw):
    return generate_password_hash(pw, method="pbkdf2:sha256")


# --------------------------------------------------------------------------
# Database helpers
# --------------------------------------------------------------------------
def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


# Seed opportunity dataset (PODS 2 real pipeline).
# Column order: deal_name, customer, assigned_am, squad, strategic_pillar,
#               estimated_value, target_quarter, stage, progress,
#               is_blocked, blocker_description, next_actions(JSON)
def _seed_deals():
    def na(items):
        return json.dumps([{"action": a, "done": d} for a, d in items])

    rows = [
        # ---------------- Anisa Rahmy ----------------
        ("Simcard IoT for DAOP (12 DAOP)", "PT Kereta Api Indonesia", "Anisa Rahmy",
         "Tender Squad", "IoT Connectivity", 600_000_000, "Q3 2026", "Negotiation", 40,
         0, "", na([("Confirm DAOP rollout scope", True), ("Finalize pricing", False)])),
        ("MDM – MS Connectivity + IoT (450 unit)", "PT Samsung Electronic Indonesia", "Anisa Rahmy",
         "Volume Squad", "Device Bundling", 2_400_000_000, "Q3 2026", "Negotiation", 50,
         0, "", na([("Align MDM spec", True), ("Commercial proposal", False)])),
        ("Tender IP Transit & Metronet Link 2", "PT Kereta Api Indonesia", "Anisa Rahmy",
         "Tender Squad", "Enterprise Solutions", 1_680_000_000, "Q3 2026", "Negotiation", 45,
         0, "", na([("Prepare tender docs", True), ("Target RFS Q3 2026", False)])),
        ("IoT Simcard – Wifi Kereta New Generation (450 unit)", "PT Kereta Api Indonesia", "Anisa Rahmy",
         "Volume Squad", "IoT Connectivity", 2_400_000_000, "Q3 2026", "Negotiation", 55,
         0, "", na([("Confirm August implementation", True), ("SIM provisioning plan", False)])),
        ("Simcard IoT for ADAS", "PT Trans Jakarta", "Anisa Rahmy",
         "Tender Squad", "IoT Connectivity", 1_250_000_000, "Q3 2026", "Prospecting", 30,
         0, "", na([("ADAS integration scoping", False), ("Target RFS Q3 2026", False)])),
        ("Connectivity HO", "GMF", "Anisa Rahmy",
         "Strategic Squad", "IoT Connectivity", 2_400_000_000, "Q3 2026", "Prospecting", 25,
         0, "", na([("HO connectivity survey", False)])),
        ("Fuel Management System", "PT Kereta Api Indonesia", "Anisa Rahmy",
         "Strategic Squad", "Enterprise Solutions", 4_000_000_000, "Q4 2026", "Prospecting", 20,
         0, "", na([("Complete POC (in progress)", False), ("2026 revenue plan", False)])),
        ("CCTV HO (Pengadaan Langsung)", "PT Trans Jakarta", "Anisa Rahmy",
         "Tender Squad", "CCTV & Vision Analytics", 200_000_000, "Q3 2026", "Prospecting", 30,
         0, "", na([("Prepare pengadaan langsung docs", False)])),

        # ---------------- Arie Prabowo ----------------
        ("Device Bundling Samsung A17 EE", "Glico Indonesia", "Arie Prabowo",
         "Volume Squad", "Device Bundling", 256_000_000, "Q3 2026", "Prospecting", 30,
         0, "", na([("Maintain fixed partner price", False), ("Confirm delivery timeline", False)])),
        ("Smart Asset Rental", "YIMM", "Arie Prabowo",
         "Strategic Squad", "Enterprise Solutions", 80_800_000, "Q3 2026", "Prospecting", 25,
         0, "", na([("Scope asset rental", False)])),
        ("Device Bundling Samsung A17 EE", "Karoseri Laksana", "Arie Prabowo",
         "Volume Squad", "Device Bundling", 1_300_000_000, "Q3 2026", "Prospecting", 35,
         0, "", na([("Lock partner price for large qty", False), ("Delivery schedule", False)])),
        ("Enterprise Asset Management", "Kobelindo Compressors", "Arie Prabowo",
         "Strategic Squad", "Enterprise Solutions", 385_000_000, "Q3 2026", "Prospecting", 30,
         0, "", na([("EAM requirement workshop", False)])),
        ("People Tracking Solution", "Triputra Agro Persada", "Arie Prabowo",
         "Strategic Squad", "Enterprise Solutions", 2_400_000_000, "Q4 2026", "Prospecting", 20,
         0, "", na([("Assign dedicated PIC", False)])),
        ("CCTV AI Crowd Analytics", "YIMM", "Arie Prabowo",
         "Strategic Squad", "CCTV & Vision Analytics", 180_000_000, "Q3 2026", "Prospecting", 30,
         0, "", na([("Frequent customer visit", False)])),
        ("CCTV Portable", "YIMM", "Arie Prabowo",
         "Volume Squad", "CCTV & Vision Analytics", 1_100_000_000, "Q3 2026", "Prospecting", 25,
         0, "", na([("Manage customer expectations", False)])),
        ("Smart Vehicle Monitoring System", "Triputra Agro Persada", "Arie Prabowo",
         "Strategic Squad", "Enterprise Solutions", 4_900_000_000, "Q4 2026", "Prospecting", 15,
         0, "", na([("Assign skilled implementation team", False)])),
        ("ESTA Vision", "Waresix", "Arie Prabowo",
         "Strategic Squad", "CCTV & Vision Analytics", 248_000_000, "Q3 2026", "Prospecting", 30,
         0, "", na([("ESTA vision demo", False)])),
        ("IoT Connectivity (Consortium)", "Consortium Mastrans Bandung", "Arie Prabowo",
         "Tender Squad", "IoT Connectivity", 1_200_000_000, "Q4 2026", "Prospecting", 20,
         0, "", na([("Consortium alignment", False)])),

        # ---------------- Ashari ----------------
        ("Strategic Account – AnterAja", "Tri Adi Bersama (AnterAja)", "Ashari",
         "Strategic Squad", "Enterprise Solutions", 7_300_000_000, "H2 2026", "Negotiation", 35,
         0, "", na([("Account plan", True), ("Solution proposal", False)])),
        ("Strategic Account – Mayora", "Cipta Niaga Semesta (Mayora Group)", "Ashari",
         "Strategic Squad", "Enterprise Solutions", 540_000_000, "H2 2026", "Prospecting", 25,
         0, "", na([("Discovery meeting", False)])),
        ("Strategic Account – IMP", "Integrasi Multi Persada (IMP)", "Ashari",
         "Strategic Squad", "Enterprise Solutions", 4_200_000_000, "H2 2026", "Prospecting", 25,
         0, "", na([("Needs assessment", False)])),
        ("Strategic Account – Cikarang Listrindo", "PT Cikarang Listrindo, Tbk.", "Ashari",
         "Strategic Squad", "Enterprise Solutions", 1_000_000_000, "H2 2026", "Prospecting", 20,
         0, "", na([("Intro meeting", False)])),
        ("Strategic Account – Indo Lysaght", "PT Indo Lysaght", "Ashari",
         "Strategic Squad", "Enterprise Solutions", 360_000_000, "H2 2026", "Prospecting", 20,
         0, "", na([("Qualify opportunity", False)])),

        # ---------------- Dimas ----------------
        ("Device Bundling – 2250 Drivers", "PT Green SM", "Dimas",
         "Volume Squad", "Device Bundling", 8_100_000_000, "Q3-Q4 2026", "Negotiation", 40,
         0, "", na([("Confirm 2250 unit rollout", True), ("Delivery scheduling", False)])),
        ("M2M IoT Simcard – Green SM Bike (10k)", "PT Green SM", "Dimas",
         "Volume Squad", "IoT Connectivity", 990_000_000, "Q3 2026", "Negotiation", 50,
         0, "", na([("August implementation plan", True), ("SIM activation", False)])),
        ("M2M IoT Simcard – Green SM EVEE Car (6k)", "PT Green SM", "Dimas",
         "Volume Squad", "IoT Connectivity", 691_000_000, "Q3 2026", "Negotiation", 45,
         0, "", na([("EVEE fleet onboarding", False)])),
        ("Device Bundling – Dexa & Ferron Pharma", "PT Dexa Group", "Dimas",
         "Volume Squad", "Device Bundling", 1_600_000_000, "Q3-Q4 2026", "Prospecting", 30,
         0, "", na([("Employee bundling scope", False)])),
        ("Digital Reward – 60TB", "PT Via Yotta Byte", "Dimas",
         "Strategic Squad", "Digital Reward", 1_900_000_000, "Q3 2026", "Negotiation", 40,
         0, "", na([("Q3 implementation", False)])),
        ("CCTV Analytics – 3 MOR Stores (30 titik)", "PT OT Group", "Dimas",
         "Strategic Squad", "CCTV & Vision Analytics", 500_000_000, "Q3 2026", "Prospecting", 30,
         0, "", na([("Site survey 3 lokasi", False), ("Installation plan 30 titik", False)])),
        ("Smart Water AI Inspection – Crystalin", "PT OT Group", "Dimas",
         "Strategic Squad", "CCTV & Vision Analytics", 332_000_000, "Q3 2026", "Prospecting", 25,
         0, "", na([("Label & coding inspection PoC", False)])),
        ("Smart Building Energy Saving", "PT Kawanlama Group", "Dimas",
         "Strategic Squad", "Enterprise Solutions", 400_000_000, "Q3 2026", "Prospecting", 25,
         0, "", na([("Energy saving assessment", False)])),
    ]
    return rows


def migrate_db(db):
    """Add columns introduced after the first release, without touching data."""
    def columns(table):
        return {r[1] for r in db.execute(f"PRAGMA table_info({table})")}

    deal_cols = columns("deals")
    if "revenue_2026" not in deal_cols:
        db.execute("ALTER TABLE deals ADD COLUMN revenue_2026 INTEGER DEFAULT 0")
        db.execute("UPDATE deals SET revenue_2026 = estimated_value WHERE revenue_2026 = 0")
    if "strategy" not in deal_cols:
        db.execute("ALTER TABLE deals ADD COLUMN strategy TEXT DEFAULT ''")
    if "proofs" not in deal_cols:
        db.execute("ALTER TABLE deals ADD COLUMN proofs TEXT DEFAULT '{}'")

    config_cols = columns("config")
    if "am_targets" not in config_cols:
        db.execute("ALTER TABLE config ADD COLUMN am_targets TEXT DEFAULT '{}'")
        db.execute("UPDATE config SET am_targets = ?", (json.dumps(DEFAULT_AM_TARGETS),))
    if "current_achievement" not in config_cols:
        db.execute("ALTER TABLE config ADD COLUMN current_achievement INTEGER DEFAULT 0")
    if "recurring_revenue" not in config_cols:
        db.execute("ALTER TABLE config ADD COLUMN recurring_revenue INTEGER DEFAULT 0")
    if "stages" not in config_cols:
        db.execute("ALTER TABLE config ADD COLUMN stages TEXT")
        db.execute("UPDATE config SET stages = ?", (json.dumps(DEFAULT_STAGES),))
    if "am_achievements" not in config_cols:
        db.execute("ALTER TABLE config ADD COLUMN am_achievements TEXT DEFAULT '{}'")
    if "am_recurring" not in config_cols:
        db.execute("ALTER TABLE config ADD COLUMN am_recurring TEXT DEFAULT '{}'")
    if "auto_stage" not in config_cols:
        # Safe to enable on upgrade: a deal with no framework progress keeps its
        # current stage (derive_stage falls back to it), so nothing is rewritten.
        db.execute("ALTER TABLE config ADD COLUMN auto_stage INTEGER DEFAULT 1")
    if "stage_rules" not in config_cols:
        db.execute("ALTER TABLE config ADD COLUMN stage_rules TEXT")
        db.execute("UPDATE config SET stage_rules = ?", (json.dumps(DEFAULT_STAGE_RULES),))
    if "max_login_logs" not in config_cols:
        db.execute("ALTER TABLE config ADD COLUMN max_login_logs INTEGER DEFAULT 100")


def init_db():
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    db.executescript(
        """
        CREATE TABLE IF NOT EXISTS deals (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            deal_name TEXT NOT NULL,
            customer TEXT DEFAULT '',
            assigned_am TEXT DEFAULT '',
            squad TEXT NOT NULL,
            strategic_pillar TEXT NOT NULL,
            estimated_value INTEGER NOT NULL,
            revenue_2026 INTEGER DEFAULT 0,
            target_quarter TEXT DEFAULT '',
            stage TEXT NOT NULL,
            progress INTEGER DEFAULT 0,
            is_blocked BOOLEAN DEFAULT 0,
            blocker_description TEXT,
            next_actions TEXT,
            strategy TEXT DEFAULT '',
            proofs TEXT DEFAULT '{}',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            role TEXT NOT NULL CHECK(role IN ('admin', 'account_manager', 'management')),
            full_name TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS login_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            username TEXT DEFAULT '',
            full_name TEXT DEFAULT '',
            role TEXT DEFAULT '',
            ip_address TEXT DEFAULT '',
            login_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS performance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            label TEXT DEFAULT '',
            source_file TEXT DEFAULT '',
            am_summary TEXT DEFAULT '[]',
            accounts TEXT DEFAULT '[]',
            uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS config (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            target_amount INTEGER DEFAULT 163000000000,
            strategic_pillars TEXT,
            squads TEXT,
            am_targets TEXT DEFAULT '{}',
            current_achievement INTEGER DEFAULT 0,
            recurring_revenue INTEGER DEFAULT 0,
            stages TEXT,
            am_achievements TEXT DEFAULT '{}',
            am_recurring TEXT DEFAULT '{}',
            auto_stage INTEGER DEFAULT 1,
            stage_rules TEXT,
            max_login_logs INTEGER DEFAULT 100,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """
    )

    # Safe, additive migrations for databases created by an earlier version.
    # ALTER only when the column is missing, so existing data is preserved.
    migrate_db(db)

    # Seed users if empty
    if db.execute("SELECT COUNT(*) FROM users").fetchone()[0] == 0:
        seed_users = [
            ("admin", "admin123", "admin", "Administrator"),
            ("exec", "exec123", "management", "Management Viewer"),
        ]
        for username, password, role, full_name in seed_users:
            db.execute(
                "INSERT INTO users (username, password, role, full_name) VALUES (?, ?, ?, ?)",
                (username, _hash(password), role, full_name),
            )
        for username, password, full_name in SEED_AMS:
            db.execute(
                "INSERT INTO users (username, password, role, full_name) VALUES (?, ?, ?, ?)",
                (username, _hash(password), "account_manager", full_name),
            )

    # Seed config if empty
    if db.execute("SELECT COUNT(*) FROM config").fetchone()[0] == 0:
        db.execute(
            """INSERT INTO config (target_amount, strategic_pillars, squads, am_targets,
                                   stages, auto_stage, stage_rules)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (163_000_000_000, json.dumps(DEFAULT_PILLARS), json.dumps(DEFAULT_SQUADS),
             json.dumps(DEFAULT_AM_TARGETS), json.dumps(DEFAULT_STAGES),
             1, json.dumps(DEFAULT_STAGE_RULES)),
        )

    # Seed deals if empty
    if db.execute("SELECT COUNT(*) FROM deals").fetchone()[0] == 0:
        db.executemany(
            """INSERT INTO deals
               (deal_name, customer, assigned_am, squad, strategic_pillar,
                estimated_value, target_quarter, stage, progress,
                is_blocked, blocker_description, next_actions)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            _seed_deals(),
        )
        # Default 2026 realizable revenue to the full TCV; admin refines per deal.
        db.execute("UPDATE deals SET revenue_2026 = estimated_value")

    db.commit()
    db.close()


# --------------------------------------------------------------------------
# Auth helpers
# --------------------------------------------------------------------------
def get_current_user():
    auth = request.headers.get("Authorization", "")
    token = auth.replace("Bearer ", "").strip() if auth else request.args.get("token", "")
    return TOKENS.get(token)


def login_required(roles=None):
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            user = get_current_user()
            if not user:
                return jsonify({"error": "Unauthorized"}), 401
            if roles and user["role"] not in roles:
                return jsonify({"error": "Forbidden"}), 403
            g.current_user = user
            return f(*args, **kwargs)

        return wrapper

    return decorator


def can_edit_deal(user, deal_row):
    """ADMIN edits anything; an AM edits only opportunities assigned to them."""
    if user["role"] == "admin":
        return True
    if user["role"] == "account_manager":
        return (deal_row["assigned_am"] or "") == (user["full_name"] or "")
    return False


def current_config(db):
    return config_to_dict(db.execute("SELECT * FROM config ORDER BY id DESC LIMIT 1").fetchone())


def resolve_stage(db, data, proofs, is_blocked, requested_stage, fallback_stage):
    """Apply the auto-derive rule unless it's off or the client asked to override."""
    cfg = current_config(db)
    if not cfg.get("auto_stage") or data.get("stage_override"):
        return requested_stage or fallback_stage
    return derive_stage(proofs, is_blocked, cfg.get("stage_rules"),
                        cfg.get("blocked_stage"), requested_stage or fallback_stage)


def deal_to_dict(row):
    return {
        "id": row["id"],
        "deal_name": row["deal_name"],
        "customer": row["customer"] or "",
        "assigned_am": row["assigned_am"] or "",
        "squad": row["squad"],
        "strategic_pillar": row["strategic_pillar"],
        "estimated_value": row["estimated_value"],
        "revenue_2026": row["revenue_2026"] if row["revenue_2026"] is not None else 0,
        "target_quarter": row["target_quarter"] or "",
        "stage": row["stage"],
        "progress": row["progress"],
        "is_blocked": bool(row["is_blocked"]),
        "blocker_description": row["blocker_description"] or "",
        "next_actions": json.loads(row["next_actions"] or "[]"),
        "strategy": (row["strategy"] if "strategy" in row.keys() else "") or "",
        "proofs": normalize_proofs(
            json.loads((row["proofs"] if "proofs" in row.keys() else "") or "{}")
        ),
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


# --------------------------------------------------------------------------
# Views
# --------------------------------------------------------------------------
@app.route("/")
def index():
    return render_template("index.html")


# --------------------------------------------------------------------------
# Auth API
# --------------------------------------------------------------------------
def client_ip():
    forwarded = request.headers.get("X-Forwarded-For", "")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.remote_addr or ""


def trim_login_logs(db, max_logs):
    """Keep only the most recent `max_logs` rows so the table never grows unbounded."""
    db.execute(
        """DELETE FROM login_logs WHERE id NOT IN (
               SELECT id FROM login_logs ORDER BY login_at DESC, id DESC LIMIT ?
           )""",
        (max_logs,),
    )


def log_login(db, row):
    """Record a successful sign-in and trim to the admin-configured retention limit."""
    db.execute(
        """INSERT INTO login_logs (user_id, username, full_name, role, ip_address)
           VALUES (?, ?, ?, ?, ?)""",
        (row["id"], row["username"], row["full_name"] or row["username"], row["role"], client_ip()),
    )
    max_logs = current_config(db).get("max_login_logs", 100)
    trim_login_logs(db, max_logs)


@app.route("/api/login", methods=["POST"])
def api_login():
    data = request.get_json(force=True) or {}
    username = data.get("username", "").strip()
    password = data.get("password", "")

    db = get_db()
    row = db.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
    if not row or not check_password_hash(row["password"], password):
        return jsonify({"error": "Invalid username or password"}), 401

    token = secrets.token_hex(24)
    TOKENS[token] = {
        "user_id": row["id"],
        "username": row["username"],
        "role": row["role"],
        "full_name": row["full_name"] or row["username"],
    }
    log_login(db, row)
    db.commit()
    return jsonify({
        "token": token,
        "role": row["role"],
        "username": row["username"],
        "full_name": row["full_name"] or row["username"],
    })


@app.route("/api/logout", methods=["POST"])
def api_logout():
    auth = request.headers.get("Authorization", "")
    token = auth.replace("Bearer ", "").strip()
    TOKENS.pop(token, None)
    return jsonify({"ok": True})


# --------------------------------------------------------------------------
# Login Logs (ADMIN only) - who signed in and when, capped at a configurable
# number of most-recent rows so the table can never grow unbounded.
# --------------------------------------------------------------------------
def login_log_to_dict(row):
    return {
        "id": row["id"],
        "username": row["username"] or "",
        "full_name": row["full_name"] or "",
        "role": row["role"] or "",
        "ip_address": row["ip_address"] or "",
        "login_at": row["login_at"],
    }


@app.route("/api/login_logs", methods=["GET"])
@login_required(roles=("admin",))
def get_login_logs():
    db = get_db()
    rows = db.execute("SELECT * FROM login_logs ORDER BY login_at DESC, id DESC").fetchall()
    return jsonify({
        "logs": [login_log_to_dict(r) for r in rows],
        "max_login_logs": current_config(db).get("max_login_logs", 100),
    })


@app.route("/api/login_logs/export_xlsx", methods=["GET"])
@login_required(roles=("admin",))
def export_login_logs_xlsx():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return jsonify({"error": "openpyxl is not installed on the server. "
                                 "Run: pip install --user openpyxl, then reload."}), 500

    db = get_db()
    rows = db.execute("SELECT * FROM login_logs ORDER BY login_at DESC, id DESC").fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Login Logs"
    ws.append(["Username", "Full Name", "Role", "Login At", "IP Address"])
    for c in range(1, 6):
        cell = ws.cell(row=1, column=c)
        cell.fill = PatternFill("solid", fgColor="1A73E8")
        cell.font = Font(color="FFFFFF", bold=True)
    ws.freeze_panes = "A2"
    for r in rows:
        ws.append([r["username"], r["full_name"], r["role"], r["login_at"], r["ip_address"]])
    for col, width in zip("ABCDE", [18, 24, 16, 20, 16]):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"login_logs_{date.today().isoformat()}.xlsx"
    return send_file(
        buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=filename)


# --------------------------------------------------------------------------
# Account Managers (for filters / assignment dropdowns) - any authenticated user
# --------------------------------------------------------------------------
@app.route("/api/deals/recalculate_stages", methods=["POST"])
@login_required(roles=("admin",))
def recalculate_stages():
    """Re-apply the auto-derive rule to every opportunity (admin, on demand)."""
    db = get_db()
    cfg = current_config(db)
    rules, blocked_stage = cfg.get("stage_rules"), cfg.get("blocked_stage")
    changed = []
    for row in db.execute("SELECT * FROM deals").fetchall():
        proofs = normalize_proofs(
            json.loads((row["proofs"] if "proofs" in row.keys() else "") or "{}"))
        new_stage = derive_stage(proofs, bool(row["is_blocked"]), rules,
                                 blocked_stage, row["stage"])
        if new_stage != row["stage"]:
            db.execute("UPDATE deals SET stage = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                       (new_stage, row["id"]))
            changed.append({"id": row["id"], "deal_name": row["deal_name"],
                            "from": row["stage"], "to": new_stage})
    db.commit()
    return jsonify({"ok": True, "changed": len(changed), "details": changed[:50]})


@app.route("/api/proof_framework", methods=["GET"])
@login_required()
def get_proof_framework():
    """The 8 Enterprise Proofs execution framework (static definition)."""
    return jsonify([{"key": k, "name": n, "description": d} for k, n, d in PROOFS])


@app.route("/api/account_managers", methods=["GET"])
@login_required()
def get_account_managers():
    db = get_db()
    rows = db.execute(
        "SELECT full_name FROM users WHERE role = 'account_manager' ORDER BY full_name"
    ).fetchall()
    names = [r["full_name"] for r in rows if r["full_name"]]
    return jsonify(names)


# --------------------------------------------------------------------------
# Deals API
# --------------------------------------------------------------------------
@app.route("/api/deals", methods=["GET"])
@login_required()
def get_deals():
    db = get_db()
    query = "SELECT * FROM deals WHERE 1=1"
    params = []
    for field, col in (("squad", "squad"), ("pillar", "strategic_pillar"),
                       ("am", "assigned_am"), ("stage", "stage"),
                       ("quarter", "target_quarter")):
        val = request.args.get(field)
        if val:
            query += f" AND {col} = ?"
            params.append(val)
    query += " ORDER BY estimated_value DESC"
    rows = db.execute(query, params).fetchall()
    return jsonify([deal_to_dict(r) for r in rows])


@app.route("/api/deals", methods=["POST"])
@login_required(roles=("admin", "account_manager"))
def create_deal():
    data = request.get_json(force=True) or {}
    user = g.current_user

    # An AM can only create opportunities assigned to themselves.
    if user["role"] == "account_manager":
        assigned_am = user["full_name"]
    else:
        assigned_am = data.get("assigned_am", "")

    est_value = int(data.get("estimated_value", 0) or 0)
    rev_2026 = data.get("revenue_2026")
    rev_2026 = int(rev_2026) if rev_2026 not in (None, "") else est_value

    db = get_db()
    new_proofs = normalize_proofs(data.get("proofs"))
    new_blocked = bool(data.get("is_blocked"))
    stage = resolve_stage(db, data, new_proofs, new_blocked,
                          data.get("stage", "Prospecting"), "Prospecting")
    cur = db.execute(
        """INSERT INTO deals
           (deal_name, customer, assigned_am, squad, strategic_pillar, estimated_value,
            revenue_2026, target_quarter, stage, progress, is_blocked, blocker_description,
            next_actions, strategy, proofs, updated_at)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)""",
        (
            data.get("deal_name", "Untitled Opportunity"),
            data.get("customer", ""),
            assigned_am,
            data.get("squad"),
            data.get("strategic_pillar"),
            est_value,
            rev_2026,
            data.get("target_quarter", ""),
            stage,
            int(data.get("progress", 0) or 0),
            1 if new_blocked else 0,
            data.get("blocker_description", ""),
            json.dumps(data.get("next_actions", [])),
            data.get("strategy", ""),
            json.dumps(new_proofs),
        ),
    )
    db.commit()
    row = db.execute("SELECT * FROM deals WHERE id = ?", (cur.lastrowid,)).fetchone()
    return jsonify(deal_to_dict(row)), 201


@app.route("/api/deals/<int:deal_id>", methods=["PUT"])
@login_required(roles=("admin", "account_manager"))
def update_deal(deal_id):
    data = request.get_json(force=True) or {}
    user = g.current_user
    db = get_db()
    row = db.execute("SELECT * FROM deals WHERE id = ?", (deal_id,)).fetchone()
    if not row:
        return jsonify({"error": "Deal not found"}), 404
    if not can_edit_deal(user, row):
        return jsonify({"error": "You can only edit opportunities assigned to you"}), 403

    # AMs cannot reassign a deal to someone else.
    if user["role"] == "account_manager":
        assigned_am = user["full_name"]
    else:
        assigned_am = data.get("assigned_am", row["assigned_am"])

    existing_strategy = row["strategy"] if "strategy" in row.keys() else ""
    existing_proofs = normalize_proofs(
        json.loads((row["proofs"] if "proofs" in row.keys() else "") or "{}")
    )
    upd_proofs = normalize_proofs(data.get("proofs", existing_proofs))
    upd_blocked = bool(data.get("is_blocked", row["is_blocked"]))
    upd_stage = resolve_stage(db, data, upd_proofs, upd_blocked,
                              data.get("stage", row["stage"]), row["stage"])
    db.execute(
        """UPDATE deals SET
             deal_name = ?, customer = ?, assigned_am = ?, squad = ?, strategic_pillar = ?,
             estimated_value = ?, revenue_2026 = ?, target_quarter = ?, stage = ?, progress = ?,
             is_blocked = ?, blocker_description = ?, next_actions = ?, strategy = ?, proofs = ?,
             updated_at = CURRENT_TIMESTAMP
           WHERE id = ?""",
        (
            data.get("deal_name", row["deal_name"]),
            data.get("customer", row["customer"]),
            assigned_am,
            data.get("squad", row["squad"]),
            data.get("strategic_pillar", row["strategic_pillar"]),
            int(data.get("estimated_value", row["estimated_value"]) or 0),
            int(data.get("revenue_2026", row["revenue_2026"] or 0) or 0),
            data.get("target_quarter", row["target_quarter"]),
            upd_stage,
            int(data.get("progress", row["progress"]) or 0),
            1 if upd_blocked else 0,
            data.get("blocker_description", row["blocker_description"]),
            json.dumps(data.get("next_actions", json.loads(row["next_actions"] or "[]"))),
            data.get("strategy", existing_strategy),
            json.dumps(upd_proofs),
            deal_id,
        ),
    )
    db.commit()
    row = db.execute("SELECT * FROM deals WHERE id = ?", (deal_id,)).fetchone()
    return jsonify(deal_to_dict(row))


@app.route("/api/deals/<int:deal_id>", methods=["DELETE"])
@login_required(roles=("admin", "account_manager"))
def delete_deal(deal_id):
    db = get_db()
    row = db.execute("SELECT * FROM deals WHERE id = ?", (deal_id,)).fetchone()
    if not row:
        return jsonify({"error": "Deal not found"}), 404
    if not can_edit_deal(g.current_user, row):
        return jsonify({"error": "You can only delete opportunities assigned to you"}), 403
    db.execute("DELETE FROM deals WHERE id = ?", (deal_id,))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/deals/<int:deal_id>/progress", methods=["PUT"])
@login_required(roles=("admin", "account_manager"))
def update_progress(deal_id):
    data = request.get_json(force=True) or {}
    progress = max(0, min(100, int(data.get("progress", 0))))
    db = get_db()
    row = db.execute("SELECT * FROM deals WHERE id = ?", (deal_id,)).fetchone()
    if not row:
        return jsonify({"error": "Deal not found"}), 404
    if not can_edit_deal(g.current_user, row):
        return jsonify({"error": "You can only edit opportunities assigned to you"}), 403
    db.execute(
        "UPDATE deals SET progress = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
        (progress, deal_id),
    )
    db.commit()
    row = db.execute("SELECT * FROM deals WHERE id = ?", (deal_id,)).fetchone()
    return jsonify(deal_to_dict(row))


@app.route("/api/deals/<int:deal_id>/blocker", methods=["PUT"])
@login_required(roles=("admin", "account_manager"))
def update_blocker(deal_id):
    data = request.get_json(force=True) or {}
    db = get_db()
    row = db.execute("SELECT * FROM deals WHERE id = ?", (deal_id,)).fetchone()
    if not row:
        return jsonify({"error": "Deal not found"}), 404
    if not can_edit_deal(g.current_user, row):
        return jsonify({"error": "You can only edit opportunities assigned to you"}), 403
    blocked = bool(data.get("is_blocked"))
    proofs = normalize_proofs(json.loads((row["proofs"] if "proofs" in row.keys() else "") or "{}"))
    stage = resolve_stage(db, data, proofs, blocked, row["stage"], row["stage"])
    db.execute(
        """UPDATE deals SET is_blocked = ?, blocker_description = ?, stage = ?,
           updated_at = CURRENT_TIMESTAMP WHERE id = ?""",
        (1 if blocked else 0, data.get("blocker_description", ""), stage, deal_id),
    )
    db.commit()
    row = db.execute("SELECT * FROM deals WHERE id = ?", (deal_id,)).fetchone()
    return jsonify(deal_to_dict(row))


# --------------------------------------------------------------------------
# Users API (ADMIN only)
# --------------------------------------------------------------------------
@app.route("/api/users", methods=["GET"])
@login_required(roles=("admin",))
def get_users():
    db = get_db()
    rows = db.execute(
        "SELECT id, username, role, full_name, created_at FROM users ORDER BY id"
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/users", methods=["POST"])
@login_required(roles=("admin",))
def create_user():
    data = request.get_json(force=True) or {}
    username = data.get("username", "").strip()
    password = data.get("password", "")
    role = data.get("role", "")
    full_name = data.get("full_name", "").strip() or username

    if not username or not password or role not in VALID_ROLES:
        return jsonify({"error": "username, password and a valid role are required"}), 400

    db = get_db()
    existing = db.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
    if existing:
        return jsonify({"error": "Username already exists"}), 409

    cur = db.execute(
        "INSERT INTO users (username, password, role, full_name) VALUES (?, ?, ?, ?)",
        (username, _hash(password), role, full_name),
    )
    db.commit()
    row = db.execute(
        "SELECT id, username, role, full_name, created_at FROM users WHERE id = ?",
        (cur.lastrowid,),
    ).fetchone()
    return jsonify(dict(row)), 201


@app.route("/api/users/<int:user_id>", methods=["PUT"])
@login_required(roles=("admin",))
def update_user(user_id):
    data = request.get_json(force=True) or {}
    db = get_db()
    row = db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    if not row:
        return jsonify({"error": "User not found"}), 404

    role = data.get("role", row["role"])
    if role not in VALID_ROLES:
        return jsonify({"error": "Invalid role"}), 400

    full_name = data.get("full_name", row["full_name"])
    password_hash = row["password"]
    if data.get("password"):
        password_hash = _hash(data["password"])

    db.execute(
        "UPDATE users SET role = ?, full_name = ?, password = ? WHERE id = ?",
        (role, full_name, password_hash, user_id),
    )
    db.commit()
    row = db.execute(
        "SELECT id, username, role, full_name, created_at FROM users WHERE id = ?",
        (user_id,),
    ).fetchone()
    return jsonify(dict(row))


@app.route("/api/users/<int:user_id>", methods=["DELETE"])
@login_required(roles=("admin",))
def delete_user(user_id):
    if g.current_user["user_id"] == user_id:
        return jsonify({"error": "Cannot delete your own account while logged in"}), 400
    db = get_db()
    db.execute("DELETE FROM users WHERE id = ?", (user_id,))
    db.commit()
    return jsonify({"ok": True})


# --------------------------------------------------------------------------
# Config API
# --------------------------------------------------------------------------
def config_to_dict(row):
    keys = row.keys()

    def jload(col, default):
        if col not in keys:
            return default
        try:
            val = json.loads(row[col] or "null")
            return default if val is None else val
        except (TypeError, ValueError):
            return default

    return {
        "target_amount": row["target_amount"],
        "strategic_pillars": json.loads(row["strategic_pillars"]),
        "squads": json.loads(row["squads"]),
        "stages": jload("stages", list(DEFAULT_STAGES)),
        "am_targets": jload("am_targets", {}),
        "am_achievements": jload("am_achievements", {}),
        "am_recurring": jload("am_recurring", {}),
        "auto_stage": bool(row["auto_stage"]) if "auto_stage" in keys and row["auto_stage"] is not None else True,
        "stage_rules": jload("stage_rules", dict(DEFAULT_STAGE_RULES)),
        "blocked_stage": DEFAULT_BLOCKED_STAGE,
        "current_achievement": (row["current_achievement"] if "current_achievement" in keys else 0) or 0,
        "recurring_revenue": (row["recurring_revenue"] if "recurring_revenue" in keys else 0) or 0,
        "max_login_logs": (row["max_login_logs"] if "max_login_logs" in keys and row["max_login_logs"] is not None else 100) or 100,
        "updated_at": row["updated_at"],
    }


@app.route("/api/config", methods=["GET"])
@login_required()
def get_config():
    db = get_db()
    row = db.execute("SELECT * FROM config ORDER BY id DESC LIMIT 1").fetchone()
    return jsonify(config_to_dict(row))


@app.route("/api/config", methods=["PUT"])
@login_required(roles=("admin",))
def update_config():
    data = request.get_json(force=True) or {}
    db = get_db()
    row = db.execute("SELECT * FROM config ORDER BY id DESC LIMIT 1").fetchone()

    current = config_to_dict(row)
    target_amount = int(data.get("target_amount", row["target_amount"]) or 0)
    strategic_pillars = data.get("strategic_pillars", json.loads(row["strategic_pillars"]))
    squads = data.get("squads", json.loads(row["squads"]))
    stages = data.get("stages", current["stages"]) or list(DEFAULT_STAGES)

    def int_map(src):
        return {k: int(v or 0) for k, v in (src or {}).items()}

    am_targets = int_map(data.get("am_targets", current["am_targets"]))
    am_achievements = int_map(data.get("am_achievements", current["am_achievements"]))
    am_recurring = int_map(data.get("am_recurring", current["am_recurring"]))
    current_achievement = int(data.get("current_achievement", current["current_achievement"]) or 0)
    recurring_revenue = int(data.get("recurring_revenue", current["recurring_revenue"]) or 0)
    auto_stage = 1 if data.get("auto_stage", current["auto_stage"]) else 0
    stage_rules = data.get("stage_rules", current["stage_rules"]) or {}
    stage_rules = {k: str(v or "") for k, v in stage_rules.items() if k in PROOF_KEYS}
    # Keep the retained login-log count within a sane range so the table can never grow unbounded.
    max_login_logs = int(data.get("max_login_logs", current["max_login_logs"]) or 100)
    max_login_logs = max(10, min(max_login_logs, 2000))

    db.execute(
        """UPDATE config SET target_amount = ?, strategic_pillars = ?, squads = ?,
           stages = ?, am_targets = ?, am_achievements = ?, am_recurring = ?,
           current_achievement = ?, recurring_revenue = ?, auto_stage = ?, stage_rules = ?,
           max_login_logs = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?""",
        (target_amount, json.dumps(strategic_pillars), json.dumps(squads),
         json.dumps(stages), json.dumps(am_targets), json.dumps(am_achievements),
         json.dumps(am_recurring), current_achievement, recurring_revenue,
         auto_stage, json.dumps(stage_rules), max_login_logs, row["id"]),
    )
    trim_login_logs(db, max_login_logs)
    db.commit()
    row = db.execute("SELECT * FROM config WHERE id = ?", (row["id"],)).fetchone()
    return jsonify(config_to_dict(row))


# --------------------------------------------------------------------------
# XLSX Backup: full export / import  (ADMIN only)
# --------------------------------------------------------------------------
DEAL_HEADERS = [
    "ID", "Opportunity", "Customer", "Account Manager", "Squad", "Strategic Pillar",
    "TCV (IDR)", "Rev 2026 (IDR)", "Target Quarter", "Stage", "Progress (%)",
    "Blocked", "Blocker", "Strategy", "Next Actions",
]
# Config keys stored as JSON (lists/dicts) vs plain integers
CONFIG_JSON_KEYS = ["strategic_pillars", "squads", "stages",
                    "am_targets", "am_achievements", "am_recurring"]
CONFIG_INT_KEYS = ["target_amount", "current_achievement", "recurring_revenue"]


def actions_to_text(actions):
    """[{action,done,due}] -> '[x] text @2026-08-15' lines (human readable + parseable)."""
    lines = []
    for a in actions or []:
        mark = "[x]" if a.get("done") else "[ ]"
        due = f" @{a['due']}" if a.get("due") else ""
        lines.append(f"{mark} {a.get('action','')}{due}")
    return "\n".join(lines)


def text_to_actions(text):
    actions = []
    for raw in str(text or "").splitlines():
        line = raw.strip()
        if not line:
            continue
        done = False
        if line.lower().startswith("[x]"):
            done, line = True, line[3:].strip()
        elif line.startswith("[ ]") or line.startswith("[]"):
            line = line.split("]", 1)[1].strip()
        due = ""
        if "@" in line:
            head, _, tail = line.rpartition("@")
            candidate = tail.strip()
            # only treat as a date if it looks like one
            if len(candidate) == 10 and candidate[4] == "-" and candidate[7] == "-":
                due, line = candidate, head.strip()
        if line:
            actions.append({"action": line, "done": done, "due": due})
    return actions


@app.route("/api/export/xlsx", methods=["GET"])
@login_required(roles=("admin",))
def export_xlsx():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return jsonify({"error": "openpyxl is not installed on the server. "
                                 "Run: pip install --user openpyxl, then reload the web app."}), 500

    db = get_db()
    deals = [deal_to_dict(r) for r in db.execute("SELECT * FROM deals ORDER BY id").fetchall()]
    config = config_to_dict(db.execute("SELECT * FROM config ORDER BY id DESC LIMIT 1").fetchone())
    users = db.execute("SELECT username, full_name, role FROM users ORDER BY id").fetchall()

    wb = Workbook()
    head_fill = PatternFill("solid", fgColor="1A73E8")
    head_font = Font(color="FFFFFF", bold=True)

    def style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill, cell.font = head_fill, head_font
        ws.freeze_panes = "A2"

    # --- Opportunities ---
    ws = wb.active
    ws.title = "Opportunities"
    ws.append(DEAL_HEADERS)
    for d in deals:
        ws.append([
            d["id"], d["deal_name"], d["customer"], d["assigned_am"], d["squad"],
            d["strategic_pillar"], d["estimated_value"], d["revenue_2026"],
            d["target_quarter"], d["stage"], d["progress"],
            "Yes" if d["is_blocked"] else "No", d["blocker_description"],
            d["strategy"], actions_to_text(d["next_actions"]),
        ])
    style_header(ws, len(DEAL_HEADERS))
    for col, width in zip("ABCDEFGHIJKLMNO",
                          [6, 38, 28, 18, 16, 22, 16, 16, 14, 14, 11, 9, 26, 50, 50]):
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        row[13].alignment = Alignment(wrap_text=True, vertical="top")  # Strategy
        row[14].alignment = Alignment(wrap_text=True, vertical="top")  # Next Actions

    # --- Execution Framework (8 Enterprise Proofs), one row per deal per proof ---
    fw = wb.create_sheet("Execution Framework")
    fw.append(["Deal ID", "Opportunity", "Account Manager", "#", "Proof",
               "Status", "Target Date", "Evidence (one per line: YYYY-MM-DD | what was done)"])
    for d in deals:
        p = d["proofs"]
        for idx, key in enumerate(PROOF_KEYS, start=1):
            item = p.get(key, {})
            evidence = "\n".join(
                (f"{e['date']} | {e['text']}" if e.get("date") else e["text"])
                for e in item.get("entries", [])
            )
            fw.append([d["id"], d["deal_name"], d["assigned_am"], idx, PROOF_NAMES[key],
                       item.get("status", "not_started"), item.get("due", ""), evidence])
    style_header(fw, 8)
    for col, width in zip("ABCDEFGH", [8, 34, 18, 5, 24, 14, 14, 52]):
        fw.column_dimensions[col].width = width
    for row in fw.iter_rows(min_row=2):
        row[7].alignment = Alignment(wrap_text=True, vertical="top")

    # --- Config ---
    cfg = wb.create_sheet("Config")
    cfg.append(["Key", "Value"])
    for k in CONFIG_INT_KEYS:
        cfg.append([k, config.get(k, 0)])
    for k in CONFIG_JSON_KEYS:
        cfg.append([k, json.dumps(config.get(k))])
    style_header(cfg, 2)
    cfg.column_dimensions["A"].width = 24
    cfg.column_dimensions["B"].width = 80

    # --- Users (no passwords are ever exported) ---
    us = wb.create_sheet("Users")
    us.append(["Username", "Full Name", "Role"])
    for u in users:
        us.append([u["username"], u["full_name"], u["role"]])
    style_header(us, 3)
    for col, width in zip("ABC", [20, 28, 20]):
        us.column_dimensions[col].width = width

    # --- Readme ---
    rm = wb.create_sheet("READ ME")
    for line in [
        ["PODS 2 Command Center - data backup"],
        [f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M')}"],
        [""],
        ["This file is both a BACKUP and an IMPORT TEMPLATE."],
        ["Re-upload it via Settings > Data Backup > Import to restore or migrate."],
        [""],
        ["Opportunities sheet:"],
        ["  - Leave ID as-is to update an existing opportunity."],
        ["  - Clear the ID to create a NEW opportunity on import."],
        ["  - Next Actions format, one per line:  [x] done action @2026-08-15"],
        ["                                        [ ] pending action"],
        ["    The @YYYY-MM-DD part is the target date and is optional."],
        ["  - Blocked column accepts Yes/No."],
        [""],
        ["Execution Framework sheet (the 8 Enterprise Proofs):"],
        ["  - One row per opportunity per proof. Keep Deal ID and Proof name unchanged."],
        ["  - Status accepts: not_started / in_progress / done / na"],
        ["  - Target Date is YYYY-MM-DD and shows up on the Calendar."],
        ["  - Evidence: one item per line. Optional date prefix, e.g."],
        ["        2026-08-01 | Workshop held with DAOP ops team"],
        ["        Budget letter received"],
        ["  - Stage is auto-derived from how far the framework has progressed"],
        ["    (configurable under Configuration > Stage automation)."],
        [""],
        ["Config sheet: JSON values - keep the JSON syntax valid."],
        ["Users sheet: passwords are never exported. New usernames on import are"],
        ["  created with the temporary password 'changeme123' - reset them right away."],
    ]:
        rm.append(line)
    rm.column_dimensions["A"].width = 95

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"deal_tracker_backup_{date.today().isoformat()}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=filename,
    )


@app.route("/api/import/xlsx", methods=["POST"])
@login_required(roles=("admin",))
def import_xlsx():
    try:
        from openpyxl import load_workbook
    except ImportError:
        return jsonify({"error": "openpyxl is not installed on the server. "
                                 "Run: pip install --user openpyxl, then reload the web app."}), 500

    upload = request.files.get("file")
    if not upload:
        return jsonify({"error": "No file uploaded"}), 400
    replace_all = str(request.form.get("replace_all", "")).lower() in ("1", "true", "yes")

    try:
        wb = load_workbook(upload, data_only=True)
    except Exception as exc:
        return jsonify({"error": f"Could not read this file as .xlsx ({exc})"}), 400

    db = get_db()
    summary = {"updated": 0, "created": 0, "deleted": 0, "users_created": 0,
               "proofs_updated": 0, "config_updated": False}

    # ---------------- Opportunities ----------------
    if "Opportunities" in wb.sheetnames:
        ws = wb["Opportunities"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        seen_ids = set()

        def s(v):
            return "" if v is None else str(v).strip()

        def n(v):
            if v is None or str(v).strip() == "":
                return 0
            try:
                return int(float(str(v).replace(".", "").replace(",", "")
                                 if isinstance(v, str) else v))
            except (TypeError, ValueError):
                return 0

        for r in rows:
            r = list(r) + [None] * (len(DEAL_HEADERS) - len(r))
            name = s(r[1])
            if not name:
                continue  # skip blank lines
            deal_id = r[0]
            payload = (
                name, s(r[2]), s(r[3]), s(r[4]), s(r[5]), n(r[6]), n(r[7]), s(r[8]),
                s(r[9]) or "Prospecting", max(0, min(100, n(r[10]))),
                1 if s(r[11]).lower() in ("yes", "true", "1") else 0,
                s(r[12]), s(r[13]), json.dumps(text_to_actions(r[14])),
            )
            existing = None
            if deal_id not in (None, ""):
                try:
                    existing = db.execute("SELECT id FROM deals WHERE id = ?",
                                          (int(deal_id),)).fetchone()
                except (TypeError, ValueError):
                    existing = None
            if existing:
                db.execute(
                    """UPDATE deals SET deal_name=?, customer=?, assigned_am=?, squad=?,
                       strategic_pillar=?, estimated_value=?, revenue_2026=?, target_quarter=?,
                       stage=?, progress=?, is_blocked=?, blocker_description=?, strategy=?,
                       next_actions=?, updated_at=CURRENT_TIMESTAMP WHERE id=?""",
                    payload + (int(deal_id),),
                )
                seen_ids.add(int(deal_id))
                summary["updated"] += 1
            else:
                cur = db.execute(
                    """INSERT INTO deals (deal_name, customer, assigned_am, squad,
                       strategic_pillar, estimated_value, revenue_2026, target_quarter, stage,
                       progress, is_blocked, blocker_description, strategy, next_actions,
                       updated_at)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,CURRENT_TIMESTAMP)""",
                    payload,
                )
                seen_ids.add(cur.lastrowid)
                summary["created"] += 1

        if replace_all and seen_ids:
            placeholders = ",".join("?" * len(seen_ids))
            cur = db.execute(f"DELETE FROM deals WHERE id NOT IN ({placeholders})",
                             tuple(seen_ids))
            summary["deleted"] = cur.rowcount

    # ---------------- Execution Framework (8 Proofs) ----------------
    if "Execution Framework" in wb.sheetnames:
        by_deal = {}
        for r in wb["Execution Framework"].iter_rows(min_row=2, values_only=True):
            r = list(r) + [None] * (8 - len(r))
            deal_id, _, _, _, proof_name, status, due, notes = r[:8]
            if deal_id in (None, "") or not proof_name:
                continue
            try:
                deal_id = int(deal_id)
            except (TypeError, ValueError):
                continue
            # match by proof name (or key), case-insensitive
            label = str(proof_name).strip().lower()
            key = next((k for k in PROOF_KEYS
                        if k == label or PROOF_NAMES[k].lower() == label), None)
            if not key:
                continue
            due_txt = ""
            if due not in (None, ""):
                due_txt = due.strftime("%Y-%m-%d") if hasattr(due, "strftime") else str(due).strip()[:10]
            # Evidence: one entry per line, optional "YYYY-MM-DD | text" prefix
            entries = []
            for line in str(notes or "").splitlines():
                line = line.strip()
                if not line:
                    continue
                date_part = ""
                if "|" in line:
                    head, _, tail = line.partition("|")
                    head = head.strip()
                    if len(head) == 10 and head[4] == "-" and head[7] == "-":
                        date_part, line = head, tail.strip()
                if line:
                    entries.append({"text": line, "date": date_part})
            by_deal.setdefault(deal_id, {})[key] = {
                "status": str(status or "not_started").strip().lower().replace(" ", "_"),
                "due": due_txt,
                "entries": entries,
            }
        for deal_id, proofs in by_deal.items():
            existing = db.execute("SELECT proofs FROM deals WHERE id = ?", (deal_id,)).fetchone()
            if not existing:
                continue
            merged = normalize_proofs(json.loads(existing["proofs"] or "{}"))
            merged.update(normalize_proofs(proofs))
            db.execute("UPDATE deals SET proofs = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                       (json.dumps(merged), deal_id))
        summary["proofs_updated"] = len(by_deal)

    # ---------------- Config ----------------
    if "Config" in wb.sheetnames:
        incoming = {}
        for k, v in wb["Config"].iter_rows(min_row=2, values_only=True):
            if not k:
                continue
            key = str(k).strip()
            if key in CONFIG_INT_KEYS:
                try:
                    incoming[key] = int(float(v or 0))
                except (TypeError, ValueError):
                    pass
            elif key in CONFIG_JSON_KEYS:
                try:
                    incoming[key] = json.loads(v) if isinstance(v, str) else v
                except (TypeError, ValueError):
                    pass
        if incoming:
            row = db.execute("SELECT * FROM config ORDER BY id DESC LIMIT 1").fetchone()
            merged = config_to_dict(row)
            merged.update(incoming)
            db.execute(
                """UPDATE config SET target_amount=?, strategic_pillars=?, squads=?, stages=?,
                   am_targets=?, am_achievements=?, am_recurring=?, current_achievement=?,
                   recurring_revenue=?, updated_at=CURRENT_TIMESTAMP WHERE id=?""",
                (int(merged["target_amount"] or 0), json.dumps(merged["strategic_pillars"]),
                 json.dumps(merged["squads"]), json.dumps(merged["stages"]),
                 json.dumps(merged["am_targets"]), json.dumps(merged["am_achievements"]),
                 json.dumps(merged["am_recurring"]), int(merged["current_achievement"] or 0),
                 int(merged["recurring_revenue"] or 0), row["id"]),
            )
            summary["config_updated"] = True

    # ---------------- Users (never overwrites existing passwords) ----------------
    if "Users" in wb.sheetnames:
        for uname, fname, role in wb["Users"].iter_rows(min_row=2, values_only=True):
            username = str(uname or "").strip()
            role = str(role or "").strip()
            if not username or role not in VALID_ROLES:
                continue
            existing = db.execute("SELECT id FROM users WHERE username = ?",
                                  (username,)).fetchone()
            if existing:
                db.execute("UPDATE users SET full_name = ?, role = ? WHERE id = ?",
                           (str(fname or "").strip() or username, role, existing["id"]))
            else:
                db.execute(
                    "INSERT INTO users (username, password, role, full_name) VALUES (?,?,?,?)",
                    (username, _hash("changeme123"), role,
                     str(fname or "").strip() or username),
                )
                summary["users_created"] += 1

    db.commit()
    return jsonify({"ok": True, "summary": summary})


# --------------------------------------------------------------------------
# Performance snapshot: import the monthly "ACH" workbook from the performance team
# --------------------------------------------------------------------------
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _num(v):
    """Excel cell -> float, tolerating blanks, text and tiny float noise."""
    if v is None or isinstance(v, bool):
        return 0.0
    if isinstance(v, (int, float)):
        return 0.0 if abs(v) < 1 else float(v)
    try:
        return float(str(v).replace(",", "").strip() or 0)
    except ValueError:
        return 0.0


def am_key(name):
    """Loose key so 'Ashari Asrar' matches the dashboard's 'Ashari'."""
    return " ".join(str(name or "").lower().replace(".", " ").split())


def am_matches(perf_name, dash_name):
    a, b = am_key(perf_name), am_key(dash_name)
    if not a or not b:
        return False
    return a == b or a.startswith(b) or b.startswith(a)


def parse_performance_workbook(wb):
    """Read the 'PODS (2)' and 'byAccount (BP)' sheets into plain dicts."""
    am_rows, accounts = [], []

    # ---- Sheet 1: per-AM monthly target/actual/forecast + MRC/pipeline/PO + summary
    sheet = None
    for nm in wb.sheetnames:
        if nm.strip().lower().startswith("pods"):
            sheet = wb[nm]
            break
    if sheet is not None:
        # Monthly triplets start at col D (index 4, 1-based); 3 cols per month.
        # Only the first block (down to its "Total" row) carries the FY summary;
        # the YTD block further down repeats the AM names with a different layout.
        for r in range(4, 40):
            first_col = str(sheet.cell(row=r, column=1).value or "").strip().lower()
            if first_col == "total":
                break
            am = sheet.cell(row=r, column=3).value
            if not am or str(am).strip().lower() in ("total", "am name"):
                continue
            monthly = []
            for mi, mname in enumerate(MONTHS):
                base = 4 + mi * 3
                monthly.append({
                    "month": mname,
                    "target": _num(sheet.cell(row=r, column=base).value),
                    # Jan-Jun are actuals, Jul-Dec are forecast in this template
                    "value": _num(sheet.cell(row=r, column=base + 1).value),
                    "is_forecast": mi >= 6,
                })
            series = lambda start: [_num(sheet.cell(row=r, column=start + i).value)
                                    for i in range(12)]
            am_rows.append({
                "am": str(am).strip(),
                "monthly": monthly,
                "mrc_monthly": series(41),        # AO..AZ
                "pipeline_monthly": series(55),   # BC..BN
                "po_monthly": series(69),         # BQ..CB
                "target_fy": _num(sheet.cell(row=r, column=85).value),   # CG
                "actual_ytd": _num(sheet.cell(row=r, column=86).value),  # CH
                "mrc_rest": _num(sheet.cell(row=r, column=87).value),    # CI
                "po_hand": _num(sheet.cell(row=r, column=88).value),     # CJ
                "forecast_fy": _num(sheet.cell(row=r, column=89).value),  # CK
                "gap": _num(sheet.cell(row=r, column=90).value),          # CL
                "conservative_pipeline": _num(sheet.cell(row=r, column=91).value),  # CM
                "current_pipeline": _num(sheet.cell(row=r, column=93).value),       # CO
            })

    # ---- Sheet 2: account-level monthly revenue
    acc_sheet = None
    for nm in wb.sheetnames:
        if "account" in nm.strip().lower():
            acc_sheet = wb[nm]
            break
    if acc_sheet is not None:
        # header row 3: month columns start at col K (11) and run while dated
        month_cols = []
        for c in range(11, acc_sheet.max_column + 1):
            h = acc_sheet.cell(row=3, column=c).value
            if hasattr(h, "strftime"):
                month_cols.append((c, h.strftime("%Y-%m")))
            elif isinstance(h, str) and h[:4].isdigit() and "-" in h:
                month_cols.append((c, h[:7]))
        for r in range(4, acc_sheet.max_row + 1):
            account = acc_sheet.cell(row=r, column=7).value
            if not account:
                continue
            months = {}
            for c, label in month_cols:
                months[label] = _num(acc_sheet.cell(row=r, column=c).value)
            if not any(months.values()):
                continue
            accounts.append({
                "pillar": str(acc_sheet.cell(row=r, column=1).value or ""),
                "revenue_category": str(acc_sheet.cell(row=r, column=4).value or ""),
                "mrc_type": str(acc_sheet.cell(row=r, column=5).value or ""),
                "account": str(account).strip(),
                "pods": str(acc_sheet.cell(row=r, column=9).value or "").strip(),
                "am": str(acc_sheet.cell(row=r, column=10).value or "").strip(),
                "months": months,
            })
    return am_rows, accounts


@app.route("/api/performance", methods=["GET"])
@login_required()
def get_performance():
    db = get_db()
    row = db.execute("SELECT * FROM performance ORDER BY id DESC LIMIT 1").fetchone()
    if not row:
        return jsonify({"available": False})
    return jsonify({
        "available": True,
        "label": row["label"],
        "source_file": row["source_file"],
        "uploaded_at": row["uploaded_at"],
        "am_summary": json.loads(row["am_summary"] or "[]"),
        "accounts": json.loads(row["accounts"] or "[]"),
    })


@app.route("/api/performance/import", methods=["POST"])
@login_required(roles=("admin",))
def import_performance():
    try:
        from openpyxl import load_workbook
    except ImportError:
        return jsonify({"error": "openpyxl is not installed on the server. "
                                 "Run: pip install --user openpyxl, then reload."}), 500

    upload = request.files.get("file")
    if not upload:
        return jsonify({"error": "No file uploaded"}), 400
    label = (request.form.get("label") or "").strip()

    try:
        wb = load_workbook(upload, data_only=True)
    except Exception as exc:
        return jsonify({"error": f"Could not read this file as .xlsx ({exc})"}), 400

    am_rows, accounts = parse_performance_workbook(wb)
    if not am_rows and not accounts:
        return jsonify({"error": "No recognisable data. Expected a sheet named like "
                                 "'PODS (2)' and one like 'byAccount (BP)'."}), 400

    db = get_db()
    db.execute(
        "INSERT INTO performance (label, source_file, am_summary, accounts) VALUES (?, ?, ?, ?)",
        (label or datetime.now().strftime("%b %Y"),
         getattr(upload, "filename", "") or "",
         json.dumps(am_rows), json.dumps(accounts)),
    )
    # keep the last 12 snapshots
    db.execute("""DELETE FROM performance WHERE id NOT IN
                  (SELECT id FROM performance ORDER BY id DESC LIMIT 12)""")
    db.commit()
    return jsonify({"ok": True, "am_count": len(am_rows), "account_rows": len(accounts)})


def _perf_am_matches(a, b):
    """Loose match: exact, or one name is a prefix of the other (mirrors the frontend)."""
    an = " ".join(str(a or "").lower().split())
    bn = " ".join(str(b or "").lower().split())
    if not an or not bn:
        return False
    return an == bn or an.startswith(bn) or bn.startswith(an)


def _perf_num(v):
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


@app.route("/api/performance/export_pdf", methods=["GET"])
@login_required()
def export_performance_pdf():
    """Performance summary as a PDF: team scorecard, gap coverage, per-AM
    breakdown, churn/growth and product-pillar mix — the same numbers shown
    on the Performance tab, snapshotted for sharing outside the dashboard."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    db = get_db()
    row = db.execute("SELECT * FROM performance ORDER BY id DESC LIMIT 1").fetchone()
    if not row:
        return jsonify({"error": "No performance data has been uploaded yet."}), 400

    am_summary = json.loads(row["am_summary"] or "[]")
    accounts = json.loads(row["accounts"] or "[]")
    am_filter = (request.args.get("am") or "").strip()

    am_rows = [a for a in am_summary if not am_filter or _perf_am_matches(a.get("am"), am_filter)]
    am_names = [a.get("am") for a in am_summary]

    def account_in_scope(r):
        if am_filter:
            return _perf_am_matches(r.get("am"), am_filter)
        return any(_perf_am_matches(r.get("am"), n) for n in am_names)

    acc_rows = [r for r in accounts if account_in_scope(r)]

    def s(key):
        return sum(_perf_num(a.get(key)) for a in am_rows)

    target, actual = s("target_fy"), s("actual_ytd")
    mrc, po, forecast, gap = s("mrc_rest"), s("po_hand"), s("forecast_fy"), s("gap")
    need, pipe = s("conservative_pipeline"), s("current_pipeline")
    attain = (actual / target * 100) if target else 0

    # ---- churn / growth: most recent two months across in-scope accounts ----
    month_keys = sorted({m for r in acc_rows for m in (r.get("months") or {}).keys()})
    by_acc = {}
    for r in acc_rows:
        key = r.get("account")
        entry = by_acc.setdefault(key, {"account": key, "am": r.get("am"), "months": {}})
        for m, v in (r.get("months") or {}).items():
            entry["months"][m] = entry["months"].get(m, 0) + _perf_num(v)
    deltas = []
    if len(month_keys) >= 2:
        cur_k, prev_k = month_keys[-1], month_keys[-2]
        for a in by_acc.values():
            cur, prev = a["months"].get(cur_k, 0), a["months"].get(prev_k, 0)
            if prev > 0 or cur > 0:
                deltas.append({"account": a["account"], "am": a["am"] or "",
                               "cur": cur, "prev": prev, "delta": cur - prev})
    down = sorted([d for d in deltas if d["delta"] < 0], key=lambda x: x["delta"])[:10]
    up = sorted([d for d in deltas if d["delta"] > 0], key=lambda x: x["delta"], reverse=True)[:10]

    # ---- revenue by product pillar (sum across all months in scope) ----
    by_pillar = {}
    for r in acc_rows:
        total = sum(_perf_num(v) for v in (r.get("months") or {}).values())
        p = r.get("pillar") or "-"
        by_pillar[p] = by_pillar.get(p, 0) + total
    pillar_entries = sorted(by_pillar.items(), key=lambda x: x[1], reverse=True)
    pillar_total = sum(v for _, v in pillar_entries) or 1

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm,
        leftMargin=1.6 * cm, rightMargin=1.6 * cm,
    )
    styles = getSampleStyleSheet()
    brand = colors.HexColor("#1a73e8")
    title_style = ParagraphStyle("PTitle", parent=styles["Title"], fontSize=18, textColor=brand)
    h2_style = ParagraphStyle("PH2", parent=styles["Heading2"], fontSize=13, textColor=brand,
                              spaceBefore=14, spaceAfter=6)
    body_style = styles["BodyText"]

    def make_table(data, col_widths):
        t = Table(data, colWidths=col_widths)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), brand),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dadce0")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f3f4")]),
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        return t

    story = [
        Paragraph("H2 2026 Command Center - PODS 2", title_style),
        Paragraph("Performance Summary Report" + (f" — {am_filter}" if am_filter else ""), styles["Heading3"]),
        Paragraph(f"Snapshot: {row['label'] or 'untitled'} (uploaded {str(row['uploaded_at'])[:10]}) "
                  f"&middot; Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", body_style),
        Spacer(1, 0.4 * cm),
    ]

    # Gap first — the top-priority number for management.
    story.append(Paragraph("Team Scorecard &amp; Gap", h2_style))
    story.append(make_table([
        ["Metric", "Value"],
        ["Target FY26", _fmt_idr(target)],
        ["Actual YTD", _fmt_idr(actual)],
        ["MRC (rest of year)", _fmt_idr(mrc)],
        ["PO on Hand", _fmt_idr(po)],
        ["Total Forecast", _fmt_idr(forecast)],
        ["Gap to Target", _fmt_idr(gap)],
        ["Attainment", f"{attain:.1f}%"],
    ], [8 * cm, 8 * cm]))
    story.append(Spacer(1, 0.3 * cm))

    story.append(Paragraph("Pipeline Cover for the Gap", h2_style))
    cover_note = (f"Gap of {_fmt_idr(gap)} needs {_fmt_idr(need)} of pipeline (conservative 3x rule). "
                  f"Current pipeline is {_fmt_idr(pipe)} — "
                  + ("covered." if pipe >= need else f"short by {_fmt_idr(max(need - pipe, 0))}."))
    story.append(Paragraph(cover_note, body_style))
    story.append(Spacer(1, 0.4 * cm))

    story.append(Paragraph("Account Manager Performance", h2_style))
    am_table_rows = [["Account Manager", "Target FY26", "Actual YTD", "% Ach", "Forecast", "Gap", "Pipeline"]]
    for a in sorted(am_rows, key=lambda x: _perf_num(x.get("target_fy")), reverse=True):
        t, ac = _perf_num(a.get("target_fy")), _perf_num(a.get("actual_ytd"))
        pct = f"{(ac / t * 100):.1f}%" if t else "-"
        am_table_rows.append([a.get("am") or "", _fmt_idr(t), _fmt_idr(ac), pct,
                              _fmt_idr(_perf_num(a.get("forecast_fy"))), _fmt_idr(_perf_num(a.get("gap"))),
                              _fmt_idr(_perf_num(a.get("current_pipeline")))])
    story.append(make_table(am_table_rows, [3.6 * cm, 2.6 * cm, 2.6 * cm, 1.6 * cm, 2.6 * cm, 2.4 * cm, 2.6 * cm]))
    story.append(Spacer(1, 0.4 * cm))

    if down or up:
        small = ParagraphStyle("psm", parent=body_style, fontSize=8)
        story.append(Paragraph("Churn Watch (vs previous month)", h2_style))
        if down:
            churn_rows = [["Account", "AM", "Previous", "Current", "Change"]]
            for d in down:
                churn_rows.append([Paragraph(d["account"], small), (d["am"] or "").split(" ")[0],
                                   _fmt_idr(d["prev"]), _fmt_idr(d["cur"]), f"-{_fmt_idr(abs(d['delta']))}"])
            story.append(make_table(churn_rows, [5.5 * cm, 2.5 * cm, 2.8 * cm, 2.8 * cm, 2.8 * cm]))
        else:
            story.append(Paragraph("No accounts declined month-over-month.", body_style))
        story.append(Spacer(1, 0.3 * cm))

        story.append(Paragraph("Growing Accounts (vs previous month)", h2_style))
        if up:
            growth_rows = [["Account", "AM", "Previous", "Current", "Change"]]
            for d in up:
                growth_rows.append([Paragraph(d["account"], small), (d["am"] or "").split(" ")[0],
                                    _fmt_idr(d["prev"]), _fmt_idr(d["cur"]), f"+{_fmt_idr(d['delta'])}"])
            story.append(make_table(growth_rows, [5.5 * cm, 2.5 * cm, 2.8 * cm, 2.8 * cm, 2.8 * cm]))
        else:
            story.append(Paragraph("No growth recorded month-over-month.", body_style))
        story.append(Spacer(1, 0.4 * cm))

    if pillar_entries:
        story.append(Paragraph("Revenue by Product Pillar", h2_style))
        pillar_rows = [["Pillar", "Revenue", "Share"]]
        for p, v in pillar_entries:
            pillar_rows.append([p, _fmt_idr(v), f"{v / pillar_total * 100:.0f}%"])
        story.append(make_table(pillar_rows, [8 * cm, 5 * cm, 3 * cm]))
        story.append(Spacer(1, 0.3 * cm))

    story.append(Paragraph(
        "For the full account-level pipeline cross-check (which top accounts have no opportunity "
        "attached in this dashboard), see the Performance tab in the live app.", body_style))

    doc.build(story)
    buf.seek(0)
    filename = f"performance_summary_{date.today().isoformat()}.pdf"
    return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name=filename)


# --------------------------------------------------------------------------
# Execution framework export in the "Sales Activity Tracker" sheet format
# --------------------------------------------------------------------------
TRACKER_HEADERS = ["No.", "PODS ", "Opportunity Name", "Customer", "Account Manager",
                   "TCV (IDR)", "Rev 2026 (IDR)", "Target Quarter",
                   "Pillar (8 Enterprise Proof)", "Activity / Action", "Status",
                   "Due Date", "Completed Date", "Notes"]
TRACKER_STATUS = {"not_started": "Not Started", "planned": "Planned", "in_progress": "In Progress",
                  "done": "Done", "na": "Not Started"}


@app.route("/api/export/tracker_xlsx", methods=["GET"])
@login_required()
def export_tracker_xlsx():
    """One row per activity, ready to paste into the shared Tracker sheet."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return jsonify({"error": "openpyxl is not installed on the server. "
                                 "Run: pip install --user openpyxl, then reload."}), 500

    db = get_db()
    query = "SELECT * FROM deals"
    params = []
    am = request.args.get("am")
    if am:
        query += " WHERE assigned_am = ?"
        params.append(am)
    query += " ORDER BY assigned_am, deal_name"
    deals = [deal_to_dict(r) for r in db.execute(query, params).fetchall()]
    pod_label = request.args.get("pod", "Pods 2")

    wb = Workbook()
    ws = wb.active
    ws.title = "Tracker"
    ws.append(TRACKER_HEADERS)
    for c in range(1, len(TRACKER_HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = PatternFill("solid", fgColor="1A73E8")
        cell.font = Font(color="FFFFFF", bold=True)
    ws.freeze_panes = "A2"

    # By default only proofs with something to say are exported (evidence, a target
    # date, or progress). ?include_empty=1 emits a row for every planned proof.
    include_empty = str(request.args.get("include_empty", "")).lower() in ("1", "true", "yes")

    n = 0
    for d in deals:
        proofs = d["proofs"]
        for idx, key in enumerate(PROOF_KEYS, start=1):
            item = proofs.get(key, {})
            status = item.get("status", "not_started")
            if status == "na":
                continue
            entries = item.get("entries", [])
            has_content = bool(entries) or bool(item.get("due")) or status != "not_started"
            if not has_content and not include_empty:
                continue
            pillar = f"{idx}. {PROOF_NAMES[key]}"
            # one row per evidence entry; a proof with none still emits its row
            rows = entries or [{"text": "", "date": "", "status": status}]
            for e in rows:
                n += 1
                e_status = e.get("status") or status
                completed = e.get("date", "") if e_status == "done" else ""
                ws.append([
                    n, pod_label, d["deal_name"], d["customer"], d["assigned_am"],
                    d["estimated_value"] or None, d["revenue_2026"] or None,
                    d["target_quarter"], pillar,
                    e.get("text") or "(activity to be defined)",
                    TRACKER_STATUS.get(e_status, "Not Started"),
                    item.get("due", ""), completed, d["strategy"],
                ])
        # ad-hoc next actions land against Proof of Qualification by default
        for a in d["next_actions"]:
            n += 1
            ws.append([
                n, pod_label, d["deal_name"], d["customer"], d["assigned_am"],
                d["estimated_value"] or None, d["revenue_2026"] or None,
                d["target_quarter"], "1. Proof of Qualification",
                a.get("action", ""), "Done" if a.get("done") else "Not Started",
                a.get("due", ""), a.get("due", "") if a.get("done") else "", d["strategy"],
            ])

    for col, width in zip("ABCDEFGHIJKLMN",
                          [6, 10, 42, 28, 24, 16, 16, 14, 26, 42, 13, 13, 15, 46]):
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        row[9].alignment = Alignment(wrap_text=True, vertical="top")
        row[13].alignment = Alignment(wrap_text=True, vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"tracker_activities_{date.today().isoformat()}.xlsx"
    return send_file(
        buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=filename)


# --------------------------------------------------------------------------
# PDF Export
# --------------------------------------------------------------------------
def _fmt_idr(value):
    if value >= 1e9:
        return f"IDR {value/1e9:,.2f}B"
    if value >= 1e6:
        return f"IDR {value/1e6:,.1f}M"
    return f"IDR {value:,.0f}"


@app.route("/api/export/pdf", methods=["GET"])
@login_required()
def export_pdf():
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    db = get_db()
    deals = [deal_to_dict(r) for r in db.execute("SELECT * FROM deals").fetchall()]
    config = config_to_dict(db.execute("SELECT * FROM config ORDER BY id DESC LIMIT 1").fetchone())

    total_pipeline = sum(d["estimated_value"] for d in deals)
    total_rev_2026 = sum(d["revenue_2026"] for d in deals)
    weighted = sum(d["estimated_value"] * d["progress"] / 100 for d in deals)
    target = config["target_amount"]
    achieved = config.get("current_achievement", 0)
    recurring = config.get("recurring_revenue", 0)
    covered = achieved + recurring + total_rev_2026
    remaining_gap = max(target - covered, 0)
    gap = target - total_pipeline
    avg_deal = total_pipeline / len(deals) if deals else 0
    am_targets = config.get("am_targets", {})

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm,
        leftMargin=1.6 * cm, rightMargin=1.6 * cm,
    )
    styles = getSampleStyleSheet()
    brand = colors.HexColor("#1a73e8")
    title_style = ParagraphStyle("TitleC", parent=styles["Title"], fontSize=18, textColor=brand)
    h2_style = ParagraphStyle("H2C", parent=styles["Heading2"], fontSize=13, textColor=brand,
                              spaceBefore=14, spaceAfter=6)
    body_style = styles["BodyText"]

    def make_table(data, col_widths):
        t = Table(data, colWidths=col_widths)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), brand),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dadce0")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f3f4")]),
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        return t

    story = []
    story.append(Paragraph("H2 2026 Command Center - PODS 2", title_style))
    story.append(Paragraph("Deal Execution &amp; Strategy Report", styles["Heading3"]))
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", body_style))
    story.append(Spacer(1, 0.4 * cm))

    story.append(Paragraph("Executive Summary", h2_style))
    story.append(Paragraph(
        f"Against the full-year 2026 target of {_fmt_idr(target)}, the team has achieved "
        f"{_fmt_idr(achieved)} to date, expects {_fmt_idr(recurring)} in recurring revenue, and is "
        f"tracking {_fmt_idr(total_rev_2026)} of realizable 2026 pipeline across {len(deals)} "
        f"opportunities. That covers {_fmt_idr(covered)}, leaving a remaining gap of "
        f"{_fmt_idr(remaining_gap)} to close. Execution spans four Account Managers across IoT "
        f"Connectivity, Device Bundling, CCTV &amp; Vision Analytics, Enterprise Solutions and "
        f"Digital Reward.",
        body_style,
    ))
    story.append(Spacer(1, 0.3 * cm))

    story.append(Paragraph("Full-Year 2026 Coverage", h2_style))
    story.append(make_table([
        ["Metric", "Value"],
        ["Full-Year 2026 Target", _fmt_idr(target)],
        ["Achieved (YTD)", _fmt_idr(achieved)],
        ["Recurring (upcoming months)", _fmt_idr(recurring)],
        ["2026 Realizable Pipeline", _fmt_idr(total_rev_2026)],
        ["Total Covered", _fmt_idr(covered)],
        ["Remaining Gap to Close", _fmt_idr(remaining_gap)],
    ], [8 * cm, 8 * cm]))
    story.append(Spacer(1, 0.4 * cm))

    story.append(Paragraph("Pipeline KPIs", h2_style))
    story.append(make_table([
        ["Metric", "Value"],
        ["Current Pipeline (TCV)", _fmt_idr(total_pipeline)],
        ["Weighted Pipeline", _fmt_idr(weighted)],
        ["Total Opportunities", str(len(deals))],
        ["Average Deal Size", _fmt_idr(avg_deal)],
    ], [8 * cm, 8 * cm]))
    story.append(Spacer(1, 0.4 * cm))

    # By Account Manager (with target attainment on 2026 revenue)
    story.append(Paragraph("Account Manager Target vs Pipeline", h2_style))
    am_val = {}
    am_rev = {}
    am_cnt = {}
    for d in deals:
        am_val[d["assigned_am"]] = am_val.get(d["assigned_am"], 0) + d["estimated_value"]
        am_rev[d["assigned_am"]] = am_rev.get(d["assigned_am"], 0) + d["revenue_2026"]
        am_cnt[d["assigned_am"]] = am_cnt.get(d["assigned_am"], 0) + 1
    am_rows = [["Account Manager", "Opps", "TCV Pipeline", "2026 Rev", "Target", "Attain."]]
    for am in sorted(am_val, key=lambda k: am_val[k], reverse=True):
        tgt = am_targets.get(am, 0)
        attain = f"{(am_rev[am]/tgt*100):.0f}%" if tgt else "-"
        am_rows.append([am, str(am_cnt[am]), _fmt_idr(am_val[am]),
                        _fmt_idr(am_rev[am]), _fmt_idr(tgt) if tgt else "-", attain])
    story.append(make_table(am_rows, [4.5 * cm, 1.3 * cm, 3.2 * cm, 3.2 * cm, 3.2 * cm, 1.6 * cm]))
    story.append(Spacer(1, 0.4 * cm))

    # By stage
    story.append(Paragraph("Deals by Stage", h2_style))
    stage_counts = {}
    for d in deals:
        stage_counts[d["stage"]] = stage_counts.get(d["stage"], 0) + 1
    story.append(make_table(
        [["Stage", "Count"]] + [[k, str(v)] for k, v in stage_counts.items()],
        [8 * cm, 8 * cm],
    ))
    story.append(Spacer(1, 0.4 * cm))

    # By pillar
    story.append(Paragraph("Pipeline by Strategic Pillar", h2_style))
    pillar_values = {}
    for d in deals:
        pillar_values[d["strategic_pillar"]] = pillar_values.get(d["strategic_pillar"], 0) + d["estimated_value"]
    story.append(make_table(
        [["Strategic Pillar", "Pipeline Value"]] +
        [[k, _fmt_idr(v)] for k, v in sorted(pillar_values.items(), key=lambda x: x[1], reverse=True)],
        [10 * cm, 6 * cm],
    ))
    story.append(Spacer(1, 0.4 * cm))

    # Opportunity detail
    story.append(Paragraph("Opportunity Detail", h2_style))
    detail_rows = [["Opportunity", "Customer", "AM", "Value", "Stage", "Prog."]]
    for d in sorted(deals, key=lambda x: x["estimated_value"], reverse=True):
        detail_rows.append([
            Paragraph(d["deal_name"], ParagraphStyle("s", parent=body_style, fontSize=7.5)),
            Paragraph(d["customer"], ParagraphStyle("s", parent=body_style, fontSize=7.5)),
            d["assigned_am"].split(" ")[0],
            _fmt_idr(d["estimated_value"]),
            d["stage"],
            f"{d['progress']}%",
        ])
    story.append(make_table(detail_rows, [5 * cm, 4 * cm, 2.2 * cm, 2.6 * cm, 2.2 * cm, 1.2 * cm]))
    story.append(Spacer(1, 0.4 * cm))

    # Execution framework coverage across the 8 Enterprise Proofs
    if deals:
        story.append(Paragraph("Execution Framework - 8 Enterprise Proofs", h2_style))
        fw_rows = [["#", "Proof", "Done", "In progress", "Not started", "N/A"]]
        for idx, key in enumerate(PROOF_KEYS, start=1):
            counts = {"done": 0, "in_progress": 0, "not_started": 0, "na": 0}
            for d in deals:
                counts[d["proofs"].get(key, {}).get("status", "not_started")] += 1
            fw_rows.append([str(idx), PROOF_NAMES[key], str(counts["done"]),
                            str(counts["in_progress"]), str(counts["not_started"]),
                            str(counts["na"])])
        story.append(make_table(fw_rows, [1 * cm, 6 * cm, 2.2 * cm, 2.6 * cm, 2.6 * cm, 1.6 * cm]))
        avg_fw = round(sum(proof_progress(d["proofs"]) for d in deals) / len(deals))
        story.append(Spacer(1, 0.2 * cm))
        story.append(Paragraph(
            f"Average framework completion across the portfolio: <b>{avg_fw}%</b>.", body_style))
        story.append(Spacer(1, 0.4 * cm))

    # Strategy playbook — the AM's how-to-close per opportunity
    strat_deals = [d for d in deals if (d.get("strategy") or "").strip()]
    if strat_deals:
        story.append(Paragraph("Strategy &amp; How to Close", h2_style))
        small = ParagraphStyle("sm", parent=body_style, fontSize=8)
        strat_rows = [["Opportunity / AM", "Strategy to win &amp; close"]]
        for d in sorted(strat_deals, key=lambda x: x["assigned_am"]):
            label = f"<b>{d['deal_name']}</b><br/>{d['customer']} &middot; {d['assigned_am']}"
            strat_rows.append([Paragraph(label, small), Paragraph(d["strategy"], small)])
        story.append(make_table(strat_rows, [6 * cm, 10 * cm]))
        story.append(Spacer(1, 0.4 * cm))

    story.append(Paragraph("Recommendations for H2 Execution", h2_style))
    blocked = [d for d in deals if d["is_blocked"]]
    recs = [
        "Prioritize Negotiation-stage opportunities with the highest weighted value to accelerate close.",
        "Protect large device-bundling deals by locking fixed partner pricing and delivery timelines.",
        "Assign dedicated, skilled implementation PICs to the largest Enterprise Solutions pursuits.",
        "Maintain frequent customer engagement and manage expectations on CCTV & Vision projects.",
    ]
    if gap > 0:
        recs.insert(0, f"Portfolio is {_fmt_idr(gap)} short of the H2 target - sustain pipeline "
                       f"generation across all four Account Managers.")
    if blocked:
        recs.insert(1, f"{len(blocked)} opportunity(ies) blocked; assign executive sponsors this week.")
    for r in recs:
        story.append(Paragraph(f"&#8226; {r}", body_style))

    doc.build(story)
    buf.seek(0)
    filename = f"deal_tracker_report_{date.today().isoformat()}.pdf"
    return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name=filename)


# --------------------------------------------------------------------------
# Bootstrap
# --------------------------------------------------------------------------
init_db()

if __name__ == "__main__":
    app.run(debug=True, port=int(os.environ.get("PORT", 5000)))
